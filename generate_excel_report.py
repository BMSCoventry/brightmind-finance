#!/usr/bin/env python3
"""
Brightmind Solutions Limited — Finance Report Generator
Generates a fully-formatted, colour-coded Excel report from the JSON export.

Usage:
    python generate_excel_report.py                    # looks for BMS_finance_data.json
    python generate_excel_report.py my_data.json       # specific input file
    python generate_excel_report.py data.json out.xlsx # specify output too
"""

import json, sys, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── Colour palette ────────────────────────────────────────────────────
C = {
    "navy":         "0F1923",
    "navy_mid":     "1E293B",
    "teal":         "0F6E56",
    "teal_dark":    "085041",
    "green_bg":     "D1FAE5",
    "green_row":    "ECFDF5",
    "green_hdr":    "059669",
    "red_bg":       "FEE2E2",
    "red_row":      "FFF5F5",
    "red_hdr":      "DC2626",
    "amber":        "D97706",
    "amber_bg":     "FFFBEB",
    "amber_hdr":    "F59E0B",
    "slate_light":  "F1F5F9",
    "slate_mid":    "CBD5E1",
    "slate_hdr":    "475569",
    "white":        "FFFFFF",
    "black":        "0F172A",
    "total_bg":     "1E293B",
    "profit_pos":   "065F46",
    "profit_neg":   "991B1B",
}

MONTHS = ['January','February','March','April','May','June',
          'July','August','September','October','November','December']
MONTHS_S = ['Jan','Feb','Mar','Apr','May','Jun',
            'Jul','Aug','Sep','Oct','Nov','Dec']

# ── Style helpers ──────────────────────────────────────────────────────
def fill(hex_code):
    return PatternFill("solid", fgColor=hex_code)

def border(color="CBD5E1", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def font(color=C["black"], bold=False, size=10, name="Calibri"):
    return Font(name=name, bold=bold, color=color, size=size)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_cell(ws, r, c, val, bg=None, fg=None, bold=False, sz=10,
             fmt=None, h_align="left", wrap=False, border_color="CBD5E1"):
    cell = ws.cell(row=r, column=c, value=val)
    if bg:   cell.fill = fill(bg)
    if fg:   cell.font = font(fg, bold, sz)
    else:    cell.font = font(C["black"], bold, sz)
    cell.alignment = align(h_align, "center", wrap)
    cell.border = border(border_color)
    if fmt:  cell.number_format = fmt
    return cell

def title_row(ws, r, cols, text, bg=C["navy_mid"], fg=C["white"], sz=13):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    cell = ws.cell(row=r, column=1, value=text)
    cell.fill = fill(bg)
    cell.font = font(fg, True, sz)
    cell.alignment = align("center", "center")
    ws.row_dimensions[r].height = 28

def section_header(ws, r, cols, text, bg, fg=C["white"]):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    cell = ws.cell(row=r, column=1, value=text)
    cell.fill = fill(bg)
    cell.font = font(fg, True, 11)
    cell.alignment = align("left", "center")
    ws.row_dimensions[r].height = 22

def col_headers(ws, r, headers, bg, fg=C["white"]):
    for c, h in enumerate(headers, 1):
        set_cell(ws, r, c, h, bg=bg, fg=fg, bold=True, sz=10,
                 h_align="center", border_color="94A3B8")
    ws.row_dimensions[r].height = 20

def total_row(ws, r, label_col_end, amount_col, total, cols,
              bg=C["total_bg"], fg=C["white"]):
    ws.merge_cells(start_row=r, start_column=1,
                   end_row=r, end_column=label_col_end)
    lbl = ws.cell(row=r, column=1, value=f"TOTAL  {ws.cell(row=r-1,column=1).value and '' or ''}")
    lbl.fill = fill(bg); lbl.font = font(fg, True, 10)
    lbl.alignment = align("right", "center")
    amt = ws.cell(row=r, column=amount_col, value=total)
    amt.fill = fill(bg); amt.font = font(fg, True, 11)
    amt.alignment = align("right", "center")
    amt.number_format = '£#,##0.00'
    for c in range(label_col_end+1, cols+1):
        if c != amount_col:
            ws.cell(row=r, column=c).fill = fill(bg)
    ws.row_dimensions[r].height = 22

# ── Month sheet ────────────────────────────────────────────────────────
def add_month_sheet(wb, year, month_idx, all_inc, all_exp):
    def in_month(d_str):
        try:
            d = datetime.fromisoformat(d_str)
            return d.month - 1 == month_idx and d.year == year
        except: return False

    m_inc = [i for i in all_inc if in_month(i.get("date",""))]
    m_exp = [e for e in all_exp if in_month(e.get("date",""))]
    total_inc = sum(float(i.get("amount",0)) for i in m_inc)
    total_exp = sum(float(e.get("amount",0)) for e in m_exp)
    net = total_inc - total_exp

    name = f"{MONTHS_S[month_idx]} {str(year)[2:]}"
    ws = wb.create_sheet(title=name)
    ws.sheet_view.showGridLines = False

    COLS = 6
    # Set column widths
    widths = [12, 26, 13, 17, 22, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    title_row(ws, r, COLS,
              f"Brightmind Solutions Limited  ·  {MONTHS[month_idx]} {year}",
              C["navy"], C["white"], 13)

    # ── Income section ──────────────────────────────────────────
    r += 2
    section_header(ws, r, COLS, "  INCOME", C["green_hdr"])
    r += 1
    col_headers(ws, r,
                ["Date","Student Name","Amount (£)","Payment Method",
                 "Payment Timing","Notes"],
                C["teal"])
    inc_start = r + 1
    for item in m_inc:
        r += 1
        bg = C["green_row"] if r % 2 == 0 else C["green_bg"]
        set_cell(ws, r, 1, item.get("date",""),          bg=bg, h_align="center")
        set_cell(ws, r, 2, item.get("student",""),       bg=bg)
        set_cell(ws, r, 3, float(item.get("amount",0)),  bg=bg, fmt='£#,##0.00', h_align="right")
        set_cell(ws, r, 4, item.get("paymentMethod","").title(), bg=bg, h_align="center")
        timing = item.get("paymentTiming","").replace("_"," ").title()
        set_cell(ws, r, 5, timing, bg=bg, h_align="center")
        set_cell(ws, r, 6, item.get("notes",""), bg=bg, wrap=True)
        ws.row_dimensions[r].height = 18
    if not m_inc:
        r += 1
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=COLS)
        ws.cell(row=r,column=1,value="No income recorded this month").fill = fill(C["green_bg"])
        ws.cell(row=r,column=1).font = font("94A3B8", False, 10)
        ws.cell(row=r,column=1).alignment = align("center","center")
    r += 1
    # Income total row
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    c1 = ws.cell(row=r,column=1,value="TOTAL INCOME")
    c1.fill=fill(C["green_hdr"]); c1.font=font(C["white"],True,11)
    c1.alignment=align("right","center"); c1.border=border("047857")
    for col in range(3,COLS+1):
        if col == 3:
            cx = ws.cell(row=r,column=col,value=total_inc)
            cx.fill=fill(C["green_hdr"]); cx.font=font(C["white"],True,12)
            cx.alignment=align("right","center"); cx.border=border("047857")
            cx.number_format='£#,##0.00'
        else:
            ws.cell(row=r,column=col).fill=fill(C["green_hdr"])
            ws.cell(row=r,column=col).border=border("047857")
    ws.row_dimensions[r].height = 24

    # ── Expenses section ─────────────────────────────────────────
    r += 2
    section_header(ws, r, COLS, "  EXPENSES", C["red_hdr"])
    r += 1
    col_headers(ws, r,
                ["Date","Description","Amount (£)","Payment Method",
                 "Category","Notes"],
                C["red_hdr"])
    for item in m_exp:
        r += 1
        bg = C["red_row"] if r % 2 == 0 else C["red_bg"]
        set_cell(ws, r, 1, item.get("date",""),          bg=bg, h_align="center")
        set_cell(ws, r, 2, item.get("description",""),   bg=bg)
        set_cell(ws, r, 3, float(item.get("amount",0)),  bg=bg, fmt='£#,##0.00', h_align="right")
        set_cell(ws, r, 4, item.get("paymentMethod","").title(), bg=bg, h_align="center")
        set_cell(ws, r, 5, item.get("category",""),      bg=bg, h_align="center")
        set_cell(ws, r, 6, item.get("notes",""),         bg=bg, wrap=True)
        ws.row_dimensions[r].height = 18
    if not m_exp:
        r += 1
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=COLS)
        ws.cell(row=r,column=1,value="No expenses recorded this month").fill = fill(C["red_bg"])
        ws.cell(row=r,column=1).font = font("94A3B8", False, 10)
        ws.cell(row=r,column=1).alignment = align("center","center")
    r += 1
    # Expense total row
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    c1 = ws.cell(row=r,column=1,value="TOTAL EXPENSES")
    c1.fill=fill(C["red_hdr"]); c1.font=font(C["white"],True,11)
    c1.alignment=align("right","center"); c1.border=border("B91C1C")
    for col in range(3,COLS+1):
        if col == 3:
            cx = ws.cell(row=r,column=col,value=total_exp)
            cx.fill=fill(C["red_hdr"]); cx.font=font(C["white"],True,12)
            cx.alignment=align("right","center"); cx.border=border("B91C1C")
            cx.number_format='£#,##0.00'
        else:
            ws.cell(row=r,column=col).fill=fill(C["red_hdr"])
            ws.cell(row=r,column=col).border=border("B91C1C")
    ws.row_dimensions[r].height = 24

    # ── Net Profit ──────────────────────────────────────────────
    r += 2
    net_color = C["profit_pos"] if net >= 0 else C["profit_neg"]
    net_bg    = "D1FAE5"          if net >= 0 else "FEE2E2"
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    lbl = ws.cell(row=r,column=1,value="NET PROFIT / (LOSS)")
    lbl.fill=fill(net_bg); lbl.font=font(net_color,True,11)
    lbl.alignment=align("right","center")
    amt = ws.cell(row=r,column=3,value=net)
    amt.fill=fill(net_bg); amt.font=font(net_color,True,13)
    amt.alignment=align("right","center"); amt.number_format='£#,##0.00'
    for col in range(4,COLS+1):
        ws.cell(row=r,column=col).fill=fill(net_bg)
    ws.row_dimensions[r].height = 26

    # Freeze top rows
    ws.freeze_panes = ws['A2']
    return ws

# ── Consolidated / FY summary ─────────────────────────────────────────
def add_consolidated_sheet(wb, fy_start, all_inc, all_exp):
    ws = wb.create_sheet(title="FY Summary")
    ws.sheet_view.showGridLines = False

    COLS = 6
    widths = [24, 16, 18, 16, 18, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    title_row(ws, r, COLS,
              f"Brightmind Solutions Limited  ·  Financial Year {fy_start}/{fy_start+1}",
              C["navy"], C["white"], 13)

    r += 2
    section_header(ws, r, COLS, "  MONTHLY INCOME & EXPENDITURE SUMMARY", C["slate_hdr"])
    r += 1
    col_headers(ws, r,
                ["Month", "Total Income (£)", "Total Expenses (£)",
                 "Net Profit (£)", "Running Total (£)", "Margin (%)"],
                C["navy_mid"])

    chart_data_row_start = r + 1
    months_labels = []
    running = 0.0
    monthly_rows = []
    for m in range(12):
        m_idx = (3 + m) % 12
        yr    = fy_start if m < 9 else fy_start + 1
        label = f"{MONTHS[m_idx]} {yr}"
        months_labels.append(MONTHS_S[m_idx])

        def in_m(d_str):
            try:
                d = datetime.fromisoformat(d_str)
                return d.month - 1 == m_idx and d.year == yr
            except: return False

        inc = sum(float(i.get("amount",0)) for i in all_inc if in_m(i.get("date","")))
        exp = sum(float(e.get("amount",0)) for e in all_exp if in_m(e.get("date","")))
        net = inc - exp
        running += net
        margin = (net / inc * 100) if inc > 0 else 0
        monthly_rows.append((label, inc, exp, net, running, margin))

    for i, row_data in enumerate(monthly_rows):
        r += 1
        bg = C["slate_light"] if i % 2 == 0 else C["white"]
        label, inc, exp, net, run, margin = row_data
        net_color  = C["profit_pos"] if net >= 0 else C["profit_neg"]
        run_color  = C["profit_pos"] if run >= 0 else C["profit_neg"]
        set_cell(ws, r, 1, label, bg=bg)
        set_cell(ws, r, 2, inc,   bg=bg, fmt='£#,##0.00', h_align="right")
        set_cell(ws, r, 3, exp,   bg=bg, fmt='£#,##0.00', h_align="right")
        c4 = set_cell(ws, r, 4, net, bg=bg, fmt='£#,##0.00', h_align="right")
        c4.font = font(net_color, True, 10)
        c5 = set_cell(ws, r, 5, run, bg=bg, fmt='£#,##0.00', h_align="right")
        c5.font = font(run_color, False, 10)
        set_cell(ws, r, 6, margin/100, bg=bg, fmt='0.0%', h_align="center")
        ws.row_dimensions[r].height = 18
    chart_data_row_end = r

    # Totals
    r += 1
    total_inc = sum(x[1] for x in monthly_rows)
    total_exp = sum(x[2] for x in monthly_rows)
    total_net = total_inc - total_exp
    total_margin = (total_net / total_inc * 100) if total_inc > 0 else 0

    def tot_cell(col, val, fmt='£#,##0.00'):
        cx = ws.cell(row=r, column=col, value=val)
        cx.fill=fill(C["total_bg"]); cx.font=font(C["white"],True,11)
        cx.alignment=align("right","center"); cx.number_format=fmt
        cx.border=border("334155")
    ws.cell(row=r,column=1,value="ANNUAL TOTAL").fill=fill(C["total_bg"])
    ws.cell(row=r,column=1).font=font(C["white"],True,11)
    ws.cell(row=r,column=1).alignment=align("left","center")
    ws.cell(row=r,column=1).border=border("334155")
    tot_cell(2, total_inc); tot_cell(3, total_exp)
    tot_cell(4, total_net); tot_cell(5, total_net)
    tot_cell(6, total_margin/100, '0.0%')
    ws.row_dimensions[r].height = 26

    # ── HMRC Self-Assessment block ────────────────────────────────────
    r += 3
    section_header(ws, r, COLS,
                   "  HMRC SELF ASSESSMENT — Schedule of Income and Expenditure",
                   C["amber"])
    r += 1
    tax_year = f"6 April {fy_start} to 5 April {fy_start+1}"
    hmrc_rows = [
        ("Tax Year",                    tax_year,       None),
        ("Business Name",               "Brightmind Solutions Limited", None),
        ("Nature of Business",          "Private Tuition / Tutoring",   None),
        ("",                            "",             None),
        ("Total Turnover (Box 15)",     total_inc,      '£#,##0.00'),
        ("Total Allowable Expenses (Box 17)", total_exp,'£#,##0.00'),
        ("Net Profit for SA103 (Box 28)",total_net,     '£#,##0.00'),
    ]
    for row_data in hmrc_rows:
        r += 1
        label, val, fmt = row_data
        bg = C["amber_bg"] if label else C["white"]
        c1 = ws.cell(row=r,column=1,value=label)
        c1.fill=fill(bg); c1.font=font(C["black"],bool(label),10)
        c1.alignment=align("left","center"); c1.border=border()
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=COLS)
        c2 = ws.cell(row=r,column=2,value=val)
        c2.fill=fill(bg); c2.border=border()
        if isinstance(val, float):
            is_neg = val < 0
            c2.font = font(C["profit_neg"] if is_neg else C["profit_pos"], True, 11)
            c2.number_format = fmt or '£#,##0.00'
            c2.alignment = align("left","center")
        else:
            c2.font=font(C["navy_mid"],False,10)
            c2.alignment=align("left","center")
        ws.row_dimensions[r].height = 20

    # ── Bar chart ─────────────────────────────────────────────────────
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = f"Monthly Income vs Expenses — FY {fy_start}/{fy_start+1}"
    chart.y_axis.title = "Amount (£)"
    chart.x_axis.title = "Month"
    chart.style = 10
    chart.width = 26
    chart.height = 14

    inc_ref  = Reference(ws, min_col=2, min_row=chart_data_row_start,
                         max_row=chart_data_row_end)
    exp_ref  = Reference(ws, min_col=3, min_row=chart_data_row_start,
                         max_row=chart_data_row_end)
    cats_ref = Reference(ws, min_col=1, min_row=chart_data_row_start,
                         max_row=chart_data_row_end)

    from openpyxl.chart import Series
    inc_series = Series(inc_ref, title="Income")
    exp_series = Series(exp_ref, title="Expenses")
    chart.series.append(inc_series)
    chart.series.append(exp_series)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, "A" + str(r + 3))

    ws.freeze_panes = ws['A2']
    return ws

# ── Expense breakdown sheet ───────────────────────────────────────────
def add_expense_breakdown(wb, fy_start, all_exp):
    ws = wb.create_sheet(title="Expense Breakdown")
    ws.sheet_view.showGridLines = False
    COLS = 4
    widths = [28, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    title_row(ws, r, COLS,
              f"Expense Breakdown by Category — FY {fy_start}/{fy_start+1}",
              C["navy"], C["white"], 13)

    # Filter to FY
    def in_fy(d_str):
        try:
            d = datetime.fromisoformat(d_str)
            fy = d.year if d.month >= 4 else d.year - 1
            return fy == fy_start
        except: return False

    fy_exp = [e for e in all_exp if in_fy(e.get("date",""))]

    # By category
    from collections import defaultdict
    cat_totals = defaultdict(float)
    cash_totals = defaultdict(float)
    card_totals = defaultdict(float)
    for e in fy_exp:
        cat = e.get("category","Other")
        amt = float(e.get("amount",0))
        cat_totals[cat] += amt
        if e.get("paymentMethod","").lower() == "cash":
            cash_totals[cat] += amt
        else:
            card_totals[cat] += amt

    r += 2
    col_headers(ws, r, ["Category","Cash (£)","Card (£)","Total (£)"], C["navy_mid"])
    grand = 0.0
    for i, (cat, total) in enumerate(sorted(cat_totals.items(),
                                             key=lambda x: -x[1])):
        r += 1
        bg = C["red_bg"] if i % 2 == 0 else C["red_row"]
        set_cell(ws, r, 1, cat, bg=bg)
        set_cell(ws, r, 2, cash_totals[cat], bg=bg, fmt='£#,##0.00', h_align="right")
        set_cell(ws, r, 3, card_totals[cat], bg=bg, fmt='£#,##0.00', h_align="right")
        set_cell(ws, r, 4, total,            bg=bg, fmt='£#,##0.00', h_align="right")
        ws.row_dimensions[r].height = 18
        grand += total

    r += 1
    for col, val in [(1,"TOTAL"),(2,sum(cash_totals.values())),(3,sum(card_totals.values())),(4,grand)]:
        cx = ws.cell(row=r,column=col,value=val)
        cx.fill=fill(C["red_hdr"]); cx.border=border("B91C1C")
        cx.font=font(C["white"],True,11)
        cx.number_format='£#,##0.00' if isinstance(val,float) else None
        cx.alignment=align("right" if col>1 else "left","center")
    ws.row_dimensions[r].height = 22

    ws.freeze_panes = ws['A2']

# ── Cash vs Bank income breakdown ─────────────────────────────────────
def add_income_breakdown(wb, fy_start, all_inc):
    ws = wb.create_sheet(title="Income Breakdown")
    ws.sheet_view.showGridLines = False
    COLS = 5
    widths = [28, 14, 14, 14, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def in_fy(d_str):
        try:
            d = datetime.fromisoformat(d_str)
            fy = d.year if d.month >= 4 else d.year - 1
            return fy == fy_start
        except: return False

    fy_inc = [i for i in all_inc if in_fy(i.get("date",""))]

    r = 1
    title_row(ws, r, COLS,
              f"Income Breakdown by Student — FY {fy_start}/{fy_start+1}",
              C["navy"], C["white"], 13)

    from collections import defaultdict
    students = defaultdict(lambda: {"cash":0.0,"bank":0.0,"total":0.0,"sessions":0})
    for i in fy_inc:
        s = i.get("student","Unknown")
        amt = float(i.get("amount",0))
        method = i.get("paymentMethod","cash").lower()
        students[s]["total"] += amt
        students[s]["sessions"] += 1
        if method == "cash":  students[s]["cash"] += amt
        else:                 students[s]["bank"] += amt

    r += 2
    col_headers(ws, r, ["Student","Cash (£)","Bank Transfer (£)",
                         "Total (£)","Sessions"], C["navy_mid"])
    for i, (name, d) in enumerate(sorted(students.items(), key=lambda x: -x[1]["total"])):
        r += 1
        bg = C["green_bg"] if i % 2 == 0 else C["green_row"]
        set_cell(ws, r, 1, name,        bg=bg)
        set_cell(ws, r, 2, d["cash"],   bg=bg, fmt='£#,##0.00', h_align="right")
        set_cell(ws, r, 3, d["bank"],   bg=bg, fmt='£#,##0.00', h_align="right")
        set_cell(ws, r, 4, d["total"],  bg=bg, fmt='£#,##0.00', h_align="right")
        set_cell(ws, r, 5, d["sessions"],bg=bg, h_align="center")
        ws.row_dimensions[r].height = 18

    r += 1
    grand = sum(d["total"] for d in students.values())
    for col, val in [(1,"TOTAL"),(2,sum(d["cash"] for d in students.values())),
                     (3,sum(d["bank"] for d in students.values())),
                     (4,grand),(5,sum(d["sessions"] for d in students.values()))]:
        cx = ws.cell(row=r,column=col,value=val)
        cx.fill=fill(C["green_hdr"]); cx.border=border("047857")
        cx.font=font(C["white"],True,11)
        cx.number_format='£#,##0.00' if isinstance(val,float) else None
        cx.alignment=align("right" if col>1 else "left","center")
    ws.row_dimensions[r].height = 22
    ws.freeze_panes = ws['A2']

# ── Main ───────────────────────────────────────────────────────────────
def generate_report(json_file, output_file=None):
    with open(json_file) as f:
        data = json.load(f)

    all_inc = data.get("income", [])
    all_exp = data.get("expenses", [])

    if not all_inc and not all_exp:
        print("⚠️  No data found in JSON file.")
        return

    # Determine financial year
    today = datetime.today()
    fy_start = today.year if today.month >= 4 else today.year - 1

    # Try to infer FY from data if available
    all_dates = [i.get("date","") for i in all_inc] + [e.get("date","") for e in all_exp]
    all_dates = [d for d in all_dates if d]
    if all_dates:
        latest = max(datetime.fromisoformat(d) for d in all_dates)
        fy_start = latest.year if latest.month >= 4 else latest.year - 1

    wb = Workbook()
    del wb["Sheet"]   # Remove default sheet

    # Add month sheets (FY: April to March)
    for m in range(12):
        m_idx = (3 + m) % 12
        yr = fy_start if m < 9 else fy_start + 1
        add_month_sheet(wb, yr, m_idx, all_inc, all_exp)

    # Add summary sheets
    add_consolidated_sheet(wb, fy_start, all_inc, all_exp)
    add_income_breakdown(wb, fy_start, all_inc)
    add_expense_breakdown(wb, fy_start, all_exp)

    if not output_file:
        output_file = f"BMS_Finance_FY{fy_start}_{fy_start+1}.xlsx"

    wb.save(output_file)
    print(f"✅  Report saved: {output_file}")
    print(f"    Sheets: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    json_input  = sys.argv[1] if len(sys.argv) > 1 else "BMS_finance_data.json"
    xlsx_output = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(json_input):
        print(f"❌  File not found: {json_input}")
        print("    Export your data from the app using 'Export JSON', then run:")
        print(f"    python generate_excel_report.py BMS_finance_data.json")
        sys.exit(1)

    generate_report(json_input, xlsx_output)
